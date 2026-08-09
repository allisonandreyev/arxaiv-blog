#!/usr/bin/env python3
"""Build static/data/arxaiv.json for the VISxAI blog post from the CVPR zip + existing maps."""
import csv, json, os, re
import numpy as np
import openpyxl

ROOT = "/Users/allisonandreyev_unrestricted/Desktop/arxaiv-vis"
ZIP = ("/private/tmp/claude-502/-Users-allisonandreyev-unrestricted-Desktop-arxaiv-vis/"
       "44e90aa7-1950-42bf-8930-edcc094ab4e8/scratchpad/cvpr/CVPR Art Gallery _26 Submission")
DATA = os.path.join(ZIP, "Data")

out = {}

# ---------------------------------------------------------------- figures map
d = json.load(open(f"{ROOT}/static/maps/both_figures.json"))
nodes, links = d["nodes"], d["links"]
byid = {n["id"]: n for n in nodes}

# per-figure diagnostic metrics (generated only)
metrics = {}
for r in csv.DictReader(open(f"{DATA}/figure_metrics.csv")):
    metrics[r["filename"]] = dict(
        clip=float(r["clip_similarity"]),
        cplx=float(r["structural_complexity"]),
        gib=float(r["gibberish_ratio"]),
        rep=float(r["repetition"]),
    )

figs = []
for n in nodes:
    gen = n["type"] == "generated"
    m = metrics.get(n["name"]) if gen else None
    f = dict(id=n["id"], n=n["name"], c=1 if gen else 0,
             x=round(n["x"], 3), y=round(n["y"], 3), z=round(n["z"], 3))
    if m:
        f.update(clip=round(m["clip"], 4), cplx=round(m["cplx"], 3),
                 gib=round(m["gib"], 4), rep=int(m["rep"]))
    figs.append(f)
out["figures"] = figs

X = np.array([[n["x"], n["y"], n["z"]] for n in nodes])
yl = np.array([0 if n["type"] == "generated" else 1 for n in nodes])


def lda_fit(X, y):
    m0, m1 = X[y == 0].mean(0), X[y == 1].mean(0)
    S = np.cov(X.T) + np.eye(X.shape[1]) * 1e-6
    w = np.linalg.solve(S, m1 - m0)
    return w, (w @ m0 + w @ m1) / 2


def lda_loo(X, y):
    c = 0
    for i in range(len(X)):
        k = np.ones(len(X), bool); k[i] = False
        w, t = lda_fit(X[k], y[k])
        c += int(w @ X[i] > t) == y[i]
    return c / len(X)


w2, t2 = lda_fit(X[:, :2], yl)
score = X[:, :2] @ w2 - t2
out["lda"] = dict(w=[float(w2[0]), float(w2[1])], t=float(t2),
                  loo2=round(float(lda_loo(X[:, :2], yl)), 4),
                  loo3=round(float(lda_loo(X, yl)), 4),
                  loo_pc1=round(float(lda_loo(X[:, :1], yl)), 4),
                  baseline=round(float(max(yl.mean(), 1 - yl.mean())), 4),
                  scores=[round(float(s), 4) for s in score])

# k-NN homophily
same = [byid[l["source"]]["type"] == byid[l["target"]]["type"] for l in links]
gen_e = [(byid[l["source"]]["type"], byid[l["target"]]["type"]) for l in links]
out["knn"] = dict(
    edges=len(links),
    same=int(sum(same)),
    same_pct=round(100 * float(np.mean(same)), 1),
    gen_pct=round(100 * np.mean([b == "generated" for a, b in gen_e if a == "generated"]), 1),
    real_pct=round(100 * np.mean([b == "real" for a, b in gen_e if a == "real"]), 1),
    k=8,
)
out["links"] = [dict(s=l["source"], t=l["target"]) for l in links]

# centroids / spread
g, r = X[yl == 0], X[yl == 1]
out["geometry"] = dict(
    cen_gen=[round(float(v), 3) for v in g.mean(0)],
    cen_real=[round(float(v), 3) for v in r.mean(0)],
    cen_dist=round(float(np.linalg.norm(g.mean(0) - r.mean(0))), 3),
    spread_gen=round(float(np.linalg.norm(g - g.mean(0), axis=1).mean()), 3),
    spread_real=round(float(np.linalg.norm(r - r.mean(0), axis=1).mean()), 3),
)

# ---------------------------------------------------------------- papers
wb = openpyxl.load_workbook(f"{DATA}/Github Release Dset.xlsx", data_only=True)
ws = wb["Sheet1"]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c else "" for c in rows[0]]
col = {h: i for i, h in enumerate(hdr) if h}
papers = {}
for row in rows[1:]:
    if not row or not row[0]:
        continue
    fn = str(row[0]).strip()
    def g(k):
        i = col.get(k)
        v = row[i] if i is not None and i < len(row) else None
        return str(v).strip() if v not in (None, "") else ""
    papers[fn] = dict(
        f=fn, title=g("Paper Title"), topic=g("Subtopic"),
        aff=g("Author Affiliations"), authors=g("Authors"),
        struct=g("Structural Integrity"),
    )
wb.close()

# NOTE: the OCR columns (G/H/I) of "Github Release Dset.xlsx" are shuffled — they were
# pasted in string-sorted filename order (1, 10, 100, 11, ...) against numerically-ordered
# rows, so paper N carries paper M's text. paper_ocr_texts.xlsx is correctly keyed by
# filename, so all text fields come from there instead.
wb = openpyxl.load_workbook(f"{DATA}/paper_ocr_texts.xlsx", data_only=True)
ws = wb.worksheets[0]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c else "" for c in rows[0]]
ci = {h: i for i, h in enumerate(hdr) if h}
for row in rows[1:]:
    if not row or not row[0]:
        continue
    fn = str(row[0]).strip()
    if fn not in papers:
        continue
    def gc(k, n):
        i = ci.get(k)
        v = row[i] if i is not None and i < len(row) else None
        return str(v).strip()[:n] if v not in (None, "", "None") else ""
    papers[fn]["abstract"] = gc("CONTENT", 1400)
    papers[fn]["refs"] = gc("REFERENCES", 900)
    s = gc("CLOSEST REAL", 300)
    papers[fn]["closest"] = re.sub(r'^s\d+\s*=\s*"?', "", s).strip().strip('"')
wb.close()

pj = json.load(open(f"{ROOT}/static/maps/papers.json"))
for n in pj["nodes"]:
    fn = n["id"]
    if fn in papers:
        papers[fn]["cluster"] = n["cluster"]
        if (not papers[fn]["title"] or papers[fn]["title"] == "None") and n.get("name"):
            papers[fn]["title"] = n["name"]

out["papers"] = sorted(papers.values(), key=lambda p: int(re.sub(r"\D", "", p["f"]) or 0))
out["paper_links"] = [dict(s=l["source"], t=l["target"], w=round(l["weight"], 4))
                      for l in pj["links"]]

# ---------------------------------------------------------------- tallies
out["clusters"] = [
    dict(id=1, name="Object detection", n=21),
    dict(id=2, name="Semantics / segmentation", n=25),
    dict(id=3, name="Image generation", n=12),
    dict(id=4, name="Localization / spatiotemporal", n=18),
    dict(id=5, name="Visual features / networks", n=24),
]
out["affiliations"] = dict(
    total_authors=340, total_affiliations=317,
    rows=[
        dict(name="Hallucinated (institution does not exist)", n=107, pct_aff=33.75, pct_auth=31.47, kind="fake"),
        dict(name="Mutated (real institution, wrong name)", n=77, pct_aff=24.29, pct_auth=22.65, kind="mutated"),
        dict(name="Stanford", n=44, pct_aff=13.88, pct_auth=12.94, kind="real"),
        dict(name="None listed", n=23, pct_aff=7.26, pct_auth=6.76, kind="none"),
        dict(name="U Toronto", n=23, pct_aff=7.26, pct_auth=6.76, kind="real"),
        dict(name="MIT", n=19, pct_aff=5.99, pct_auth=5.59, kind="real"),
        dict(name="UC Berkeley", n=16, pct_aff=5.05, pct_auth=4.71, kind="real"),
        dict(name="Google / DeepMind", n=15, pct_aff=4.73, pct_auth=4.41, kind="real"),
        dict(name="CMU", n=10, pct_aff=3.15, pct_auth=2.94, kind="real"),
        dict(name="Georgia Tech", n=7, pct_aff=2.21, pct_auth=2.06, kind="real"),
        dict(name="UIUC", n=6, pct_aff=1.89, pct_auth=1.76, kind="real"),
    ])
out["structure"] = [
    dict(name="All features present", pct=69),
    dict(name="Missing title", pct=19),
    dict(name="Missing references", pct=8),
    dict(name="Missing affiliations", pct=7),
]
out["names"] = dict(
    last=[("Chen", 28), ("Zhang", 17), ("Wang", 11), ("Liu", 11), ("Lee", 10),
          ("Thompson", 9), ("Johnson", 9), ("Nguyen", 8), ("Smith", 8), ("Patel", 8),
          ("Doe", 7), ("Wu", 6), ("Wong", 5), ("Li", 5), ("Kim", 5), ("Xu", 5)],
    first=[("Michael", 20), ("David", 17), ("Emily", 13), ("John", 13), ("Kevin", 11),
           ("Laura", 9), ("Rachel", 8), ("Alex", 7), ("Daniel", 6), ("Eric", 6),
           ("Jane", 6), ("Jason", 5), ("Thomas", 5), ("Alice", 5), ("James", 5)],
    full=[("Kevin Chen", 4), ("David Chen", 4), ("John Doe", 4), ("Jacob Wang", 3),
          ("Emily Chen", 3), ("John Smith", 3), ("Jane Doe", 3), ("Alice Johnson", 3),
          ("Jane Smith", 3)],
)
out["clip_sim"] = dict(g2r=0.44537696, g2g=0.49628806, r2r=0.4854793)

# metric summary
gm = [f for f in figs if f["c"] == 1]
def summ(key):
    v = np.array([f[key] for f in gm])
    return dict(mean=round(float(v.mean()), 4), sd=round(float(v.std(ddof=1)), 4),
                min=round(float(v.min()), 4), max=round(float(v.max()), 4),
                med=round(float(np.median(v)), 4))
out["metric_summary"] = {k: summ(k) for k in ("clip", "cplx", "gib", "rep")}
out["counts"] = dict(papers=len(out["papers"]), gen_figs=int((yl == 0).sum()),
                     real_figs=int((yl == 1).sum()))

# ------------------------------------------------- pixel-statistics ablation
# Seven cheap appearance statistics, recomputed here for BOTH sets, so the post can
# test whether the CLIP separation is explainable by low-level looks alone.
from PIL import Image
Image.MAX_IMAGE_PIXELS = None
PKEYS = ["edge", "ink", "colf", "sat", "ent", "white", "ar"]


def pixfeats(path):
    im = Image.open(path).convert("RGB")
    im.thumbnail((256, 256), Image.LANCZOS)
    a = np.asarray(im).astype(np.float32) / 255.0
    gray = a @ np.array([0.299, 0.587, 0.114], np.float32)
    gx = np.abs(np.diff(gray, axis=1))[:-1, :]
    gy = np.abs(np.diff(gray, axis=0))[:, :-1]
    R, G, B = a[..., 0], a[..., 1], a[..., 2]
    rg, yb = R - G, 0.5 * (R + G) - B
    mx, mn = a.max(-1), a.min(-1)
    h, _ = np.histogram(gray, bins=64, range=(0, 1))
    h = h / h.sum()
    return dict(
        edge=float(np.sqrt(gx ** 2 + gy ** 2).mean()),
        ink=float((gray < 0.85).mean()),
        colf=float(np.sqrt(rg.std() ** 2 + yb.std() ** 2)
                   + 0.3 * np.sqrt(rg.mean() ** 2 + yb.mean() ** 2)),
        sat=float(np.where(mx > 0, (mx - mn) / np.maximum(mx, 1e-6), 0).mean()),
        ent=float(-(h[h > 0] * np.log2(h[h > 0])).sum()),
        white=float((gray > 0.95).mean()),
        ar=float(im.width / im.height),
    )


P = []
for f in figs:
    dirn = "figures" if f["c"] == 1 else "real-figures"
    P.append(pixfeats(f"{ROOT}/static/images/{dirn}/{f['n']}"))
for f, pf in zip(figs, P):
    f["p"] = {k: round(pf[k], 5) for k in PKEYS}

Xp = np.array([[pf[k] for k in PKEYS] for pf in P])
Xp = (Xp - Xp.mean(0)) / Xp.std(0)
yr = (yl == 1).astype(int)          # 1 = real


def auc(pos, neg):
    allv = np.concatenate([pos, neg])
    order = allv.argsort()
    ranks = np.empty(len(allv)); ranks[order] = np.arange(1, len(allv) + 1)
    n1, n2 = len(pos), len(neg)
    return (ranks[:n1].sum() - n1 * (n1 + 1) / 2) / (n1 * n2)


PLABEL = dict(edge="Edge density", ink="Ink coverage", colf="Colourfulness",
              sat="Saturation", ent="Tone entropy", white="White space",
              ar="Aspect ratio")
out["pixel"] = dict(
    keys=PKEYS, labels=PLABEL,
    rows=[dict(k=k,
               label=PLABEL[k],
               ai=round(float(Xp[yl == 0][:, i].mean()), 4),
               real=round(float(Xp[yl == 1][:, i].mean()), 4),
               ai_raw=round(float(np.array([pf[k] for pf, c in zip(P, yl) if c == 0]).mean()), 4),
               real_raw=round(float(np.array([pf[k] for pf, c in zip(P, yl) if c == 1]).mean()), 4),
               auc=round(float(max(auc(Xp[yl == 0][:, i], Xp[yl == 1][:, i]),
                                   1 - auc(Xp[yl == 0][:, i], Xp[yl == 1][:, i]))), 3),
               loo=round(float(lda_loo(Xp[:, [i]], yr)), 4))
          for i, k in enumerate(PKEYS)],
)
out["ablation"] = [
    dict(name="Always guess “generated”", acc=round(float(max(yr.mean(), 1 - yr.mean())), 4), kind="none"),
    dict(name="7 pixel statistics", acc=round(float(lda_loo(Xp, yr)), 4), kind="pixel"),
    dict(name="CLIP PC₁ only", acc=round(float(lda_loo(X[:, :1], yr)), 4), kind="clip"),
    dict(name="CLIP PC₁ + PC₂", acc=round(float(lda_loo(X[:, :2], yr)), 4), kind="clip"),
    dict(name="CLIP PCs + pixel statistics", acc=round(float(lda_loo(np.hstack([X[:, :2], Xp]), yr)), 4), kind="both"),
]

os.makedirs(f"{ROOT}/static/data", exist_ok=True)
with open(f"{ROOT}/static/data/arxaiv.json", "w") as fh:
    json.dump(out, fh, separators=(",", ":"))
print("wrote", os.path.getsize(f"{ROOT}/static/data/arxaiv.json"), "bytes")
print("papers:", len(out["papers"]), "figs:", len(figs))
print("lda:", {k: v for k, v in out["lda"].items() if k != "scores"})
print("knn:", out["knn"])
print("metric summary:", json.dumps(out["metric_summary"], indent=1))
miss = [p["f"] for p in out["papers"] if not p.get("closest")]
print("papers missing closest-real:", len(miss))
print("sample paper:", json.dumps(out["papers"][0])[:600])
